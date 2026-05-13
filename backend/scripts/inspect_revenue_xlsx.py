import openpyxl

xlsx_path = 'D:/workspace/caiwu04/doc/收入毛利明细-脱敏测试用.xlsx'

wb = openpyxl.load_workbook(xlsx_path, data_only=False)
print(f"Sheets: {wb.sheetnames}")

# Check second sheet
ws2 = wb.worksheets[1]
print(f"\nSecond sheet dimensions: {ws2.dimensions}")
for row_idx in range(1, min(10, ws2.max_row + 1)):
    for col_idx in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            print(f"  Row {row_idx}, Col {col_idx}: {repr(cell.value)}")

# Check first sheet more carefully - maybe data is stored differently
ws1 = wb.worksheets[0]
print(f"\n=== First sheet - scanning ALL cols for non-None (rows 2-100) ===")
for col_idx in range(1, ws1.max_column + 1):
    header = ws1.cell(row=1, column=col_idx).value
    for row_idx in range(2, min(101, ws1.max_row + 1)):
        cell = ws1.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            print(f"  Col {col_idx-1} ({repr(header)[:25]}): row {row_idx} has {repr(cell.value)[:40]}")
            break

wb.close()

# Also check the file size
import os
size = os.path.getsize(xlsx_path)
print(f"\nFile size: {size:,} bytes ({size/1024:.1f} KB)")
