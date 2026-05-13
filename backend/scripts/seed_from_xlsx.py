"""
Seed financial_data from 收入毛利明细-脱敏测试用.xlsx
Uses existing org/period columns + mock financial metric values.
"""
import asyncio
import os
import random
from pathlib import Path

import openpyxl
import asyncpg

XLSX_PATH = Path(__file__).parent.parent.parent / "doc" / "收入毛利明细-脱敏测试用.xlsx"


async def seed():
    print("Reading xlsx with openpyxl...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"\nKey columns: 0=确认收入日期, 8=销售BGBU, 9=产品线")

    # Col indices (1-based in xlsx, 0-based in Python)
    # xlsx col 46 (index 45) = 收入金额(人民币), xlsx col 47 (index 46) = 不含税成本
    COL_PERIOD = 0    # 确认收入日期 (YYYY-MM format)
    COL_BGBU = 8       # 销售BGBU
    COL_PRODUCT = 9    # 产品线
    COL_REVENUE = 45   # 收入金额(人民币) - xlsx col 46
    COL_COST = 46      # 不含税成本 - xlsx col 47

    print("\nBuilding records...")
    records = []

    row_count = 0
    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]

        period = str(vals[COL_PERIOD])[:7] if vals[COL_PERIOD] else "2026-03"
        company = str(vals[COL_BGBU]) if vals[COL_BGBU] else "DEFAULT"
        product = str(vals[COL_PRODUCT]) if vals[COL_PRODUCT] else "其他"

        # Read actual revenue and cost from xlsx columns
        raw_revenue = vals[COL_REVENUE]
        raw_cost = vals[COL_COST]

        # Convert to float, default to 0 if None/invalid
        try:
            base_revenue = float(raw_revenue) if raw_revenue is not None else 0.0
        except (ValueError, TypeError):
            base_revenue = 0.0
        try:
            cost = float(raw_cost) if raw_cost is not None else 0.0
        except (ValueError, TypeError):
            cost = 0.0

        # Generate realistic mock values when xlsx columns are empty
        if base_revenue == 0.0 and cost == 0.0:
            base_revenue = abs(random.gauss(50000, 15000))
            cost = base_revenue * random.uniform(0.55, 0.75)

        profit = base_revenue - cost
        margin = (profit / base_revenue * 100) if base_revenue > 0 else 0.0

        records.append(('revenue', round(base_revenue, 2), period, company, None, None))
        records.append(('cost', round(cost, 2), period, company, None, None))
        records.append(('gross_profit', round(profit, 2), period, company, None, None))
        records.append(('achievement_rate', round(random.uniform(0.85, 1.05), 4), period, company, None, None))
        records.append(('profit_margin', round(margin, 2), period, company, None, None))

        row_count += 1
        if row_count % 1000 == 0:
            print(f"  Processed {row_count:,} rows, records: {len(records):,}")

    wb.close()
    print(f"\nTotal rows: {row_count:,}, Total records: {len(records):,}")

    if not records:
        print("ERROR: No records generated!")
        return

    database_url = os.environ.get(
        "DATABASE_URL",
        "postgresql://learnhouse:learnhouse@localhost:5432/caiwu",
    )
    conn = await asyncpg.connect(database_url)
    await conn.execute("TRUNCATE TABLE financial_data RESTART IDENTITY CASCADE")
    print("Cleared financial_data")

    batch_size = 50_000
    total_inserted = 0
    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        await conn.copy_records_to_table(
            "financial_data",
            columns=["metric_name", "metric_value", "period", "entity", "tags", "raw_row"],
            records=batch,
        )
        total_inserted += len(batch)
        print(f"  Inserted {total_inserted:,}/{len(records):,}")

    await conn.close()
    print(f"\nDone! Total: {total_inserted:,}")

    conn2 = await asyncpg.connect(database_url)
    count = await conn2.fetchval("SELECT COUNT(*) FROM financial_data")
    periods = await conn2.fetch("SELECT DISTINCT period FROM financial_data ORDER BY period LIMIT 5")
    metrics = await conn2.fetch(
        "SELECT metric_name, COUNT(*), SUM(metric_value) FROM financial_data "
        "WHERE metric_name IN ('revenue','cost','gross_profit') GROUP BY metric_name"
    )
    print(f"\nVerification:")
    print(f"  Total rows: {count:,}")
    print(f"  Periods: {[p[0] for p in periods]}")
    for m in metrics:
        print(f"  {m[0]}: count={m[1]:,}, sum={m[2]:,.2f}")
    await conn2.close()


if __name__ == "__main__":
    asyncio.run(seed())
